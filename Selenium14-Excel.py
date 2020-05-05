import openpyxl

workbook = openpyxl.load_workbook("data1.xlsx")

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column
print(rows,cols)

for r in range(1,rows+1):
	for c in range(1,cols+1):
		print(sheet.cell(row=r,column=c).value,end = "     ")
	print()

# Writing Data into Excel

workbookNew = openpyxl.load_workbook("data2.xlsx")

sheetNew = workbookNew.active

for r in range(1,rows+1):
	for c in range(1,cols+1):
		sheetNew.cell(row=r,column=c).value = sheet.cell(row=r,column=c).value

workbookNew.save("data2.xlsx")
