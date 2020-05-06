import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

driver = webdriver.Chrome("/Users/shreyas_rl/Desktop/git/Selenium-Python/chromedriver")

driver.get("http://www.mca.gov.in/mcafoportal/showVerifyDIN.do")

def getRowCount(file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_row)

def getColCount(file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_column)


def readFileData(file,sheetName,rowNum,colNum):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.cell(row=rowNum,column=colNum).value)
	

def writeDataIntoFile(file,sheetName,rowNum,colNum,data):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	sheet.cell(row=rowNum,column=colNum).value = data
	workbook.save(file)


path = "data.xlsx"

rows = getRowCount(path,'Sheet1')
cols = getColCount(path,'Sheet1')

print(rows,cols)

for r in range(24,rows+1):
	DIN = readFileData(path,'Sheet1',r,1)
	driver.find_element_by_id("DIN").send_keys(DIN)
	driver.find_element_by_id("verifyDIN_0").click()
	time.sleep(1)
	try:
		"""driver.find_element_by_id("directorFullName").is_displayed():
								driver.find_element_by_id("msgboxclose")
								writeDataIntoFile(path,'Sheet1',r,2,'Not Approved')"""

		dirName = driver.find_element_by_id("directorFullName").get_attribute("value")
	except:
		dirName = 'Not Approved'
		driver.find_element_by_id("msgboxclose").click()
	finally:
		print(str(r) + ":" + dirName)
		writeDataIntoFile(path,'Sheet1',r,2,dirName)
	driver.get("http://www.mca.gov.in/mcafoportal/showVerifyDIN.do")


driver.quit()

