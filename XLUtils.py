import openpyxl


def getRowCount(file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_row)

def getColCount(file,sheetName):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.max_column)


def readFileData(file,sheetName,rowNum,colNum):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	return(sheet.cell(row=rowNum,column=colNum).value)
	

def writeDataIntoFile(file,sheetName,rowNum,colNum,data):
	workbook = openpyxl.load_workbook(file)
	sheet = workbook.get_sheet_by_name(sheetName)
	sheet.cell(row=rowNum,column=colNum).value = data
	workbook.save(file)

